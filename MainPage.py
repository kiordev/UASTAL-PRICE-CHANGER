import flet as ft
import openpyxl
import os
from design import *
from flet import FilePicker, FilePickerResultEvent

uastal_orange = "#E67E22"  # Цвет из design.py (можешь подставить свой)

def update_prices(import_path, target_path):
    import_wb = openpyxl.load_workbook(import_path)
    target_wb = openpyxl.load_workbook(target_path)

    import_ws = import_wb.active
    target_ws = target_wb.active

    # Считываем артикулы и цены из шаблона импорта
    price_dict = {}
    for row in import_ws.iter_rows(min_row=2, max_col=3, values_only=True):
        article, retail, special = row
        if article:
            price_dict[str(article).strip()] = (retail, special)

    # Обновляем прайс
    for row in target_ws.iter_rows(min_row=8, min_col=10, max_col=10):  # Колонка J
        for cell in row:
            article = str(cell.value).strip() if cell.value else None
            if article in price_dict:
                retail_price, special_price = price_dict[article]
                target_ws[f'H{cell.row}'] = retail_price
                target_ws[f'I{cell.row}'] = special_price

    # Сохраняем новый файл
    new_path = target_path.replace(".xlsx", "_UPDATED.xlsx")
    target_wb.save(new_path)
    return new_path


def main(page: ft.Page):
    page.title = "UASTAL PRICE CHANGER"
    page.window.height = 200
    page.window.width = 500
    page.window.resizable = False
    page.window.maximizable = False
    page.padding = 0
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER

    import_file_path = None
    target_file_path = None

    def pick_import_file_result(e: FilePickerResultEvent):
        nonlocal import_file_path
        if e.files:
            import_file_path = e.files[0].path
            importButton.text = "ТАБЛИЦЯ +"
            importButton.bgcolor = ft.Colors.GREEN
            page.update()

    def pick_target_file_result(e: FilePickerResultEvent):
        nonlocal target_file_path
        if e.files:
            target_file_path = e.files[0].path
            targetButton.text = "ПРАЙС +"
            targetButton.bgcolor = ft.Colors.GREEN
            page.update()

    def do_import(e):
        if import_file_path and target_file_path:
            try:
                new_file = update_prices(import_file_path, target_file_path)
                page.snack_bar = ft.SnackBar(ft.Text(f"✅ Файл оновлено: {os.path.basename(new_file)}"))
                page.snack_bar.open = True
                page.update()
            except Exception as ex:
                page.snack_bar = ft.SnackBar(ft.Text(f"❌ Помилка: {ex}"))
                page.snack_bar.open = True
                page.update()
        else:
            page.snack_bar = ft.SnackBar(ft.Text("⚠️ Оберіть обидва файли!"))
            page.snack_bar.open = True
            page.update()

    # Виджеты
    mainText = ft.Text(
        "UASTAL PRICE IMPORT",
        weight=ft.FontWeight.W_700,
        size=30,
        color=uastal_orange,
    )

    importButton = ft.FilledButton(
        "ТАБЛИЦЯ ІМПОРТУ",
        color=ft.colors.WHITE,
        bgcolor=uastal_orange,
        width=180,
        height=40,
    )

    targetButton = ft.FilledButton(
        "СТАРИЙ ПРАЙС",
        color=ft.colors.WHITE,
        bgcolor=uastal_orange,
        width=180,
        height=40,
    )

    makeImportButton = ft.FilledButton(
        "ОНОВИТИ ЦІНИ",
        color=ft.colors.WHITE,
        bgcolor=uastal_orange,
        width=180,
        height=40,
    )

    # File pickers
    import_picker = FilePicker(on_result=pick_import_file_result)
    target_picker = FilePicker(on_result=pick_target_file_result)
    page.overlay.extend([import_picker, target_picker])

    importButton.on_click = lambda e: import_picker.pick_files(allow_multiple=False)
    targetButton.on_click = lambda e: target_picker.pick_files(allow_multiple=False)
    makeImportButton.on_click = do_import

    # Разметка
    buttonRow = ft.Row([
        importButton, 
        targetButton],
        alignment=ft.MainAxisAlignment.SPACE_AROUND)
    
    mainColumn = ft.Column([
        mainText, 
        buttonRow, 
        makeImportButton],
        alignment=ft.MainAxisAlignment.START,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER)

    mainContainer = ft.Container(
        content=mainColumn,
        gradient=ft.LinearGradient(begin=ft.alignment.bottom_right, end=ft.alignment.top_left, colors=[ft.Colors.DEEP_ORANGE_50, ft.Colors.ORANGE_50]),
        height=600,
        width=1000,
    )

    page.add(mainContainer)

ft.app(target=main)
